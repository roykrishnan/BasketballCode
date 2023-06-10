import time, openpyxl
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import pandas as pd 
import numpy as np

# Set up. Picking our scraping tools & methods. 
driver_path = '/path/to/webdriver/executable'
service = Service(driver_path)
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # Run in headless mode (without opening browser window)
driver = webdriver.Chrome(service=service, options=options)

# Load the webpage, setting wait time to let data load & getting HTML.
# input = ('enter URL of scrape: ')
url = 'https://2kleague.nba.com/game/1212300009/HVGRUG/'
driver.get(url)
time.sleep(5)  
html = driver.page_source

# Create a BeautifulSoup object to parse the HTML content. Getting quarters. 
soup = BeautifulSoup(html, 'html.parser')
combined_elements = soup.find_all(class_=['q1', 'quarter'])
#Is.digit strips everything in the frame that is not a number.
combined_texts = [element.get_text(strip=True) for element in combined_elements if element.get_text(strip=True).isdigit()]
# Store filtered elements in quarters:
away_team = combined_texts[:4]
home_team = combined_texts[5:9]
# Find the second table on the page
tables = soup.find_all('table')
if len(tables) >= 2:
    second_table = tables[1]  # Assuming the second table is at index 1
    # Extract the table content
    table_rows = second_table.find_all('tr')
    
    # Iterate over the table rows and print the data, excluding the first and last rows
    row_data_list = []
    for i, row in enumerate(table_rows):
        if i == 0 or i == len(table_rows) - 1:
            continue
        row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
        row_data_list.append(row_data)

    away_pg = row_data_list[0]
    away_sg = row_data_list[1]
    away_sf = row_data_list[2]
    away_pf = row_data_list[3]
    away_c = row_data_list[4]

if len(tables) >= 3:
    third_table = tables[2]  # Assuming the third table is at index 2
    # Extract the table content
    table_rows = third_table.find_all('tr')

    # Iterate over the table rows and print the data, excluding the first and last rows
    home_row_data_list = []
    for i, row in enumerate(table_rows):
        if i == 0 or i == len(table_rows) - 1:
            continue
        row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
        home_row_data_list.append(row_data)

    home_pg = home_row_data_list[0]
    home_sg = home_row_data_list[1]
    home_sf = home_row_data_list[2]
    home_pf = home_row_data_list[3]
    home_c = home_row_data_list[4]

# Create dataframes to be sorted and then read into Excel file. 
df = pd.DataFrame({
    "PG": home_pg,
    "SG": home_sg,
    "SF": home_sf,
    "PF": home_pf,
    "C": home_c
})

df2 = pd.DataFrame({
    "PG": away_pg,
    "SG": away_sg,
    "SF": away_sf,
    "PF": away_pf,
    "C": away_c
})

# Transpose the DataFrames
df = df.transpose()
df2 = df2.transpose()

# Save the transposed DataFrames to the Excel file. First row are variable names.
with pd.ExcelWriter('Game1.xlsx') as writer:
    df.to_excel(writer, sheet_name="home team", index=False)
    df2.to_excel(writer, sheet_name="away team", index=False)
    print ("Player Data has been scraped and sorted into Excel")

#Open file, delete first row. Replace with: var_index = ['PTS','REB', 'AST', 'STL', 'BLK', 'TOV', 'FGM/A', '3PM/A', 'FTM/A']. Close file. 
# Open the workbook
workbook = load_workbook('/Users/rohitkrishnan/Desktop/Game1.xlsx')
# Get the sheet names
sheet_names = workbook.sheetnames
# Iterate over the sheets
for sheet_name in sheet_names[:2]:
    # Get the sheet by name
    sheet = workbook[sheet_name]
    # Define the new header
    new_header = ['Name','PTS', 'REB', 'AST', 'STL', 'BLK', 'PF','TOV', 'FGM/A', '3PM/A', 'FTM/A']
    # Replace the values in the first row with the new header
    for col_idx, header_value in enumerate(new_header, start=1):
        sheet.cell(row=1, column=col_idx).value = header_value
    
# Adding quarter values from before here. To be appended to to the home/away sheets.
quarters = ['Q1', 'Q2', 'Q3', 'Q4']
for sheet_name in ['home team', 'away team']:
    # Select the specific sheet
    sheet = workbook[sheet_name]
    # Find the last row in the sheet
    last_row = sheet.max_row + 1
    # Print the header list as a row before the data
    for index, value in enumerate(quarters):
        sheet.cell(row=last_row, column=index+1).value = value
# Home Quarters
sheet = workbook['home team']
last_row = sheet.max_row + 1
for index, value in enumerate(home_team):
    sheet.cell(row=last_row, column=index+1).value = value
# Away Quarters
sheet = workbook['away team']
last_row = sheet.max_row + 1
for index, value in enumerate(away_team):
    sheet.cell(row=last_row, column=index+1).value = value

# Save the modified workbook, overwriting the old file
workbook.save('/Users/rohitkrishnan/Desktop/Game1.xlsx')
print ("File is now updated")

# Close the webdriver
driver.quit()
